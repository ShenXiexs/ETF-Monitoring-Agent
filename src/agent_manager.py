from __future__ import annotations

import hashlib
import http.client
import io
import json
import os
import re
import sqlite3
import ssl
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional
from xml.sax.saxutils import escape

from dotenv import load_dotenv
from openpyxl import load_workbook

try:
    from .financial_skills import DOCUMENT_WORKFLOWS, MODULE_SKILLS, build_skillbook, get_module_skill_cards, get_skill_cards
except ImportError:
    from financial_skills import DOCUMENT_WORKFLOWS, MODULE_SKILLS, build_skillbook, get_module_skill_cards, get_skill_cards

BASE_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BASE_DIR / ".env"
DEMO_BUNDLE_PATH = BASE_DIR / "data" / "demo_workspace.json"
if ENV_PATH.exists():
    load_dotenv(ENV_PATH, override=True)
else:
    load_dotenv()

MODULES = [
    {
        "key": "product_research",
        "label": "产品研究",
        "summary": "回答产品属性、规模、代码、成立信息与结构化基础事实。",
        "prompt": "你是资管产品工作台中的产品研究模块，负责提供准确、克制、结构化的产品事实说明。",
    },
    {
        "key": "market_monitoring",
        "label": "市场监测",
        "summary": "跟踪成交活跃度、资金流向、重点指数波动和日内异动。",
        "prompt": "你是资管产品工作台中的市场监测模块，负责解释行情信号、成交变化与资金流向。",
    },
    {
        "key": "content_strategy",
        "label": "内容策略",
        "summary": "将市场事实整理成传播卖点、节奏建议与内容框架。",
        "prompt": "你是资管产品工作台中的内容策略模块，负责将事实整理为可执行的沟通框架与节奏建议。",
    },
    {
        "key": "policy_analysis",
        "label": "政策解析",
        "summary": "处理监管动态、制度文件摘要与政策研判报告。",
        "prompt": "你是资管产品工作台中的政策解析模块，负责解读政策文件、识别影响并给出中性判断。",
    },
]

class AssetWorkbenchManager:
    def __init__(
        self,
        data_source_dir: Optional[str] = None,
        profile_path: Optional[str] = None,
        enable_demo_mode: bool = True,
    ) -> None:
        self.base_dir = BASE_DIR
        self.data_source_dir = Path(data_source_dir).expanduser() if data_source_dir else self._resolve_data_source_dir()
        self.profile_path = Path(profile_path).expanduser() if profile_path else None
        self.enable_demo_mode = enable_demo_mode
        self.data: Dict[str, dict] = {}
        self.dates: List[str] = []
        self.current_index = 0
        self.policies: List[dict] = []
        self.all_signals: Dict[str, List[dict]] = {}
        self.warnings: List[str] = []
        self.mode = "empty"
        self.profile = self._load_profile()
        self.cache_path = Path(tempfile.gettempdir()) / "asset_intel_workbench_llm_cache.db"
        self._init_cache()
        self._load_external_data()
        self.vector_store = None
        if self.dates:
            self._precalculate_all_signals()
            self._init_vector_store()

    def _resolve_data_source_dir(self) -> Optional[Path]:
        raw = os.getenv("DATA_SOURCE_DIR", "").strip()
        return Path(raw).expanduser() if raw else None

    def _profile_default(self) -> dict:
        default_path = self.base_dir / "config" / "default_profile.json"
        return json.loads(default_path.read_text(encoding="utf-8"))

    def _profile_path(self) -> Optional[Path]:
        if self.profile_path:
            return self.profile_path
        raw = os.getenv("DATA_PROFILE_PATH", "").strip()
        return Path(raw).expanduser() if raw else None

    def _merge_dicts(self, base: dict, override: dict) -> dict:
        merged = dict(base)
        for key, value in override.items():
            if isinstance(value, dict) and isinstance(merged.get(key), dict):
                merged[key] = self._merge_dicts(merged[key], value)
            else:
                merged[key] = value
        return merged

    def _load_profile(self) -> dict:
        profile = self._profile_default()
        custom_path = self._profile_path()
        if custom_path:
            if custom_path.exists():
                custom = json.loads(custom_path.read_text(encoding="utf-8"))
                profile = self._merge_dicts(profile, custom)
            else:
                self.warnings = getattr(self, "warnings", [])
                self.warnings.append(f"未找到 DATA_PROFILE_PATH 指定的配置文件：{custom_path}")
        return profile

    def _init_cache(self) -> None:
        conn = sqlite3.connect(self.cache_path)
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS cache (
                key_hash TEXT PRIMARY KEY,
                prompt TEXT,
                response TEXT,
                model TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        conn.commit()
        conn.close()

    def _get_cache(self, key_hash: str) -> Optional[str]:
        try:
            conn = sqlite3.connect(self.cache_path)
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT response
                FROM cache
                WHERE key_hash = ?
                  AND timestamp > datetime('now', '-1 day')
                """,
                (key_hash,),
            )
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else None
        except Exception:
            return None

    def _set_cache(self, key_hash: str, prompt: str, response: str, model: str) -> None:
        try:
            conn = sqlite3.connect(self.cache_path)
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT OR REPLACE INTO cache (key_hash, prompt, response, model, timestamp)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                """,
                (key_hash, prompt, response, model),
            )
            conn.commit()
            conn.close()
        except Exception:
            return

    def _load_external_data(self) -> None:
        existing_warnings = list(self.warnings)
        self.data = {}
        self.policies = []
        self.dates = []
        self.current_index = 0
        self.mode = "empty"
        self.warnings = existing_warnings

        if not self.data_source_dir:
            if self._load_demo_bundle("未配置 DATA_SOURCE_DIR，已自动切换到内置演示模式。"):
                return
            self.warnings.append("未配置 DATA_SOURCE_DIR，工作台以空态模式启动。")
            return

        if not self.data_source_dir.exists():
            if self._load_demo_bundle(f"未找到外部数据目录：{self.data_source_dir}，已自动切换到内置演示模式。"):
                return
            self.warnings.append(f"未找到外部数据目录：{self.data_source_dir}")
            return

        files = self.profile.get("files", {})
        snapshot_path = self.data_source_dir / files.get("market_snapshot", "market_snapshot.json")
        policy_path = self.data_source_dir / files.get("policy_catalog", "policy_catalog.xlsx")

        if snapshot_path.exists():
            self.data = self._load_market_snapshot(snapshot_path)
            self.dates = sorted(self.data.keys())
        else:
            self.warnings.append(f"缺少文件：{snapshot_path.name}")

        if policy_path.exists():
            self.policies = self._load_policies(policy_path)
        else:
            self.warnings.append(f"缺少文件：{policy_path.name}")

        if not self.data:
            if self._load_demo_bundle("当前未加载任何可用市场快照，已自动切换到内置演示模式。"):
                return
            self.warnings.append("当前未加载任何市场快照，监测与日报区域将显示空态。")
            return

        self.mode = "external"

    def _load_demo_bundle(self, reason: str) -> bool:
        if not self.enable_demo_mode or not DEMO_BUNDLE_PATH.exists():
            return False
        try:
            bundle = json.loads(DEMO_BUNDLE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return False

        snapshot = bundle.get("snapshot", {})
        policies = bundle.get("policies", [])
        normalized = self._normalize_snapshot_bundle(snapshot)
        if not normalized:
            return False

        self.data = normalized
        self.dates = sorted(self.data.keys())
        self.policies = self._normalize_policy_rows(policies)
        self.mode = "demo"
        self.warnings.append(reason)
        self.warnings.append("当前使用内置脱敏演示数据，可直接加载 Demo Case 完成比赛展示。")
        return True

    def _load_market_snapshot(self, snapshot_path: Path) -> Dict[str, dict]:
        raw = json.loads(snapshot_path.read_text(encoding="utf-8"))
        return self._normalize_snapshot_bundle(raw)

    def _normalize_snapshot_bundle(self, raw: object) -> Dict[str, dict]:
        normalized: Dict[str, dict] = {}
        snapshot_profile = self.profile.get("snapshot", {})
        date_field = snapshot_profile.get("date_field", "date")
        products_field = snapshot_profile.get("products_field", "products")
        indices_field = snapshot_profile.get("indices_field", "indices")

        if isinstance(raw, list):
            iterable = []
            for item in raw:
                if isinstance(item, dict) and item.get(date_field):
                    iterable.append((item[date_field], item))
        elif isinstance(raw, dict):
            iterable = raw.items()
        else:
            iterable = []

        for date_key, payload in iterable:
            date_str = self._normalize_date(str(date_key))
            if not date_str or not isinstance(payload, dict):
                continue

            products = [
                item
                for item in (self._normalize_product(entry) for entry in payload.get(products_field, []))
                if item
            ]
            indices = self._normalize_indices(payload.get(indices_field, {}))
            normalized[date_str] = {"products": products, "indices": indices}

        return normalized

    def _load_policies(self, policy_path: Path) -> List[dict]:
        policy_profile = self.profile.get("policy", {})
        workbook = load_workbook(policy_path, data_only=True, read_only=True)
        sheet_name = policy_profile.get("sheet_name")
        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        headers = [cell.value for cell in sheet[1]]
        header_lookup = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}
        date_idx = self._match_header(header_lookup, policy_profile.get("columns", {}).get("date", ["公告日期"]))
        title_idx = self._match_header(header_lookup, policy_profile.get("columns", {}).get("title", ["标题"]))
        rank_idx = self._match_header(header_lookup, policy_profile.get("columns", {}).get("rank", ["法律位阶"]))
        source_idx = self._match_header(header_lookup, policy_profile.get("columns", {}).get("source", ["来源"]))
        policies: List[dict] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            record = list(row)
            policies.append(
                {
                    "公告日期": self._normalize_date(record[date_idx]) if date_idx is not None and date_idx < len(record) else None,
                    "标题": str(record[title_idx] if title_idx is not None and title_idx < len(record) else "").strip(),
                    "法律位阶": str(
                        (record[rank_idx] if rank_idx is not None and rank_idx < len(record) else "")
                        or policy_profile.get("default_rank", "行业规定")
                    ).strip(),
                    "来源": str(record[source_idx] if source_idx is not None and source_idx < len(record) else "").strip(),
                }
            )
        return self._normalize_policy_rows(policies)

    def _normalize_policy_rows(self, rows: List[dict]) -> List[dict]:
        normalized: List[dict] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            record = {
                "公告日期": self._normalize_date(row.get("公告日期") or row.get("date") or row.get("日期")),
                "标题": str(row.get("标题") or row.get("title") or "").strip(),
                "法律位阶": str(
                    row.get("法律位阶") or row.get("rank") or row.get("分类") or self.profile.get("policy", {}).get("default_rank", "行业规定")
                ).strip(),
                "来源": str(row.get("来源") or row.get("source") or row.get("链接") or "").strip(),
            }
            if record["标题"]:
                normalized.append(record)
        return normalized

    def _match_header(self, header_lookup: Dict[str, int], aliases: List[str]) -> Optional[int]:
        for alias in aliases:
            if alias in header_lookup:
                return header_lookup[alias]
        lowered_lookup = {key.lower(): value for key, value in header_lookup.items()}
        for alias in aliases:
            if alias.lower() in lowered_lookup:
                return lowered_lookup[alias.lower()]
        return None

    def _normalize_indices(self, raw_indices: object) -> Dict[str, dict]:
        normalized: Dict[str, dict] = {}
        index_fields = self.profile.get("snapshot", {}).get("index_fields", {})
        code_field = index_fields.get("code", "code")
        if isinstance(raw_indices, list):
            iterator = []
            for item in raw_indices:
                if isinstance(item, dict) and item.get(code_field):
                    iterator.append((item[code_field], item))
        elif isinstance(raw_indices, dict):
            iterator = raw_indices.items()
        else:
            iterator = []

        for code, payload in iterator:
            if not isinstance(payload, dict):
                continue
            normalized[str(code)] = {
                "name": str(payload.get(index_fields.get("name", "name"), "")).strip(),
                "prev_close": self._safe_float(payload.get(index_fields.get("prev_close", "prev_close"))),
                "open": self._safe_float(payload.get(index_fields.get("open", "open"))),
                "change": self._safe_float(payload.get(index_fields.get("change", "change"))),
                "volume": self._safe_float(payload.get(index_fields.get("volume", "volume"))),
            }
        return normalized

    def _normalize_product(self, payload: object) -> Optional[dict]:
        if not isinstance(payload, dict):
            return None
        product_fields = self.profile.get("snapshot", {}).get("product_fields", {})
        code = str(payload.get(product_fields.get("code", "code"), "")).strip()
        name = str(payload.get(product_fields.get("name", "name"), "")).strip()
        if not code or not name:
            return None
        return {
            "code": code,
            "name": name,
            "setup_date": self._normalize_date(payload.get(product_fields.get("setup_date", "setup_date"))),
            "list_date": self._normalize_date(payload.get(product_fields.get("list_date", "list_date"))),
            "scale": self._safe_float(payload.get(product_fields.get("scale", "scale"))),
            "volume": self._safe_float(payload.get(product_fields.get("volume", "volume"))),
            "inflow": self._safe_float(payload.get(product_fields.get("inflow", "inflow"))),
            "index_code": str(payload.get(product_fields.get("index_code", "index_code"), "")).strip(),
        }

    def _normalize_date(self, value: object) -> Optional[str]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        match = re.match(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
        if match:
            return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
        return None

    @staticmethod
    def _safe_float(value: object) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            text = str(value).strip().replace(",", "")
            if text in {"", "--", "-"}:
                return 0.0
            return float(text)
        except Exception:
            return 0.0

    def _precalculate_all_signals(self) -> None:
        original_index = self.current_index
        for idx, _ in enumerate(self.dates):
            self.current_index = idx
            self.all_signals[self.dates[idx]] = self.check_rules()
        self.current_index = original_index

    def _init_vector_store(self) -> None:
        try:
            try:
                from .vector_store import get_vector_store
            except ImportError:
                from vector_store import get_vector_store
            self.vector_store = get_vector_store()
            self.vector_store.build_signal_index(self.all_signals)
        except Exception:
            self.vector_store = None

    def has_data(self) -> bool:
        return bool(self.dates)

    def reset_cycle(self) -> None:
        self.current_index = 0

    def next_step(self) -> bool:
        if not self.has_data():
            return False
        if self.current_index < len(self.dates) - 1:
            self.current_index += 1
            return True
        return False

    def get_current_date(self) -> Optional[str]:
        state = self.get_current_state()
        return state["date"] if state else None

    def get_current_state(self) -> Optional[dict]:
        if not self.has_data() or self.current_index >= len(self.dates):
            return None
        current_date = self.dates[self.current_index]
        return {
            "date": current_date,
            "products": self.data[current_date].get("products", []),
            "indices": self.data[current_date].get("indices", {}),
            "policies": [item for item in self.policies if item.get("公告日期") == current_date],
        }

    def get_all_products(self) -> List[dict]:
        seen = set()
        products: List[dict] = []
        for date in self.dates:
            for item in self.data[date].get("products", []):
                if item["code"] in seen:
                    continue
                seen.add(item["code"])
                products.append({"code": item["code"], "name": item["name"]})
        return sorted(products, key=lambda item: item["name"])

    def get_market_leaderboard(self, category: str = "volume", top_n: int = 3) -> List[dict]:
        state = self.get_current_state()
        if not state:
            return []
        if category not in {"volume", "inflow", "scale"}:
            category = "volume"
        return sorted(state["products"], key=lambda item: item.get(category, 0), reverse=True)[:top_n]

    def get_daily_summary(self) -> str:
        state = self.get_current_state()
        if not state:
            return "当前未加载外部市场快照。"

        date = state["date"]
        summary_lines = [f"【{date} 市场摘要】"]
        signals = self.all_signals.get(date, [])
        highlights = [item["content"] for item in signals if item["category"] in {"规模里程碑", "行情异动"}][:3]
        if highlights:
            summary_lines.append("- 重点信号：" + "；".join(highlights))

        indices = state.get("indices", {})
        index_lines = []
        for item in indices.values():
            change = item.get("change", 0)
            if abs(change) < 0.8:
                continue
            direction = "上涨" if change > 0 else "下跌"
            index_lines.append(f"{item['name']} {direction}{abs(change):.1f}%")
        if index_lines:
            summary_lines.append("- 指数表现：" + "，".join(index_lines[:3]))
        return "\n".join(summary_lines)

    def get_date_summary(self, date_str: str) -> str:
        if date_str not in self.data:
            return f"未找到 {date_str} 的快照数据。"

        day_data = self.data[date_str]
        signals = self.all_signals.get(date_str, [])
        lines = [f"### {date_str} 数据摘要", ""]

        if signals:
            lines.append("**当日信号**")
            lines.extend(f"- [{item['category']}] {item['content']}" for item in signals[:5])
            lines.append("")

        top_volume = sorted(day_data.get("products", []), key=lambda item: item.get("volume", 0), reverse=True)[:3]
        if top_volume:
            lines.append("**成交活跃 Top 3**")
            lines.extend(
                f"{idx}. {item['name']} ({item['code']}) 成交额 {item['volume']:.2f} 亿"
                for idx, item in enumerate(top_volume, 1)
            )
            lines.append("")

        top_inflow = sorted(day_data.get("products", []), key=lambda item: item.get("inflow", 0), reverse=True)[:3]
        if top_inflow:
            lines.append("**资金流入 Top 3**")
            lines.extend(
                f"{idx}. {item['name']} ({item['code']}) 净流入 {item['inflow']:.2f} 亿"
                for idx, item in enumerate(top_inflow, 1)
            )

        return "\n".join(lines).strip()

    def get_date_range_comparison(self, start_date: str, end_date: str, product_code: Optional[str] = None) -> str:
        dates_in_range = [date for date in self.dates if start_date <= date <= end_date]
        if not dates_in_range:
            return f"未找到 {start_date} 至 {end_date} 的数据。"

        if product_code:
            rows = []
            for date in dates_in_range:
                product = self.find_product(product_code, date)
                if product:
                    rows.append(product)
            if not rows:
                return f"区间内未找到产品 {product_code}。"

            lines = [
                f"### {product_code} 区间对比",
                "",
                "| 日期 | 规模(亿) | 成交额(亿) | 净流入(亿) |",
                "|:---|---:|---:|---:|",
            ]
            for idx, date in enumerate(dates_in_range):
                product = self.find_product(product_code, date)
                if not product:
                    continue
                lines.append(
                    f"| {date} | {product['scale']:.2f} | {product['volume']:.2f} | {product['inflow']:.2f} |"
                )
            return "\n".join(lines)

        total_inflow = 0.0
        total_volume = 0.0
        highlights: List[str] = []
        for date in dates_in_range:
            products = self.data[date].get("products", [])
            total_inflow += sum(item.get("inflow", 0) for item in products)
            total_volume += sum(item.get("volume", 0) for item in products)
            highlights.extend(item["content"] for item in self.all_signals.get(date, [])[:2])

        lines = [
            f"### {start_date} 至 {end_date} 区间对比",
            "",
            f"- 累计净流入：{total_inflow:.2f} 亿",
            f"- 累计成交额：{total_volume:.2f} 亿",
        ]
        if highlights:
            lines.append("- 重点信号：")
            lines.extend(f"  - {item}" for item in highlights[:6])
        return "\n".join(lines)

    def empty_daily_report(self) -> dict:
        return {
            "available": False,
            "date": None,
            "highlights": [],
            "news": [],
            "suggestions": ["当前未加载外部数据，日报内容暂不可用。"],
        }

    def get_daily_report(self, date_str: Optional[str]) -> dict:
        if not date_str or date_str not in self.data:
            return self.empty_daily_report()

        state = self.data[date_str]
        products = state.get("products", [])
        signals = self.all_signals.get(date_str, [])
        highlights = []

        for category, metric, unit in (
            ("规模领先", "scale", "亿"),
            ("成交活跃", "volume", "亿"),
            ("资金流入", "inflow", "亿"),
        ):
            ranked = sorted(products, key=lambda item: item.get(metric, 0), reverse=True)
            if ranked:
                top_item = ranked[0]
                highlights.append(
                    {
                        "category": category,
                        "product_name": top_item["name"],
                        "product_code": top_item["code"],
                        "value": top_item.get(metric, 0),
                        "unit": unit,
                    }
                )

        news = [item["content"] for item in signals[:6]]
        suggestions = []
        if highlights:
            suggestions.append(f"围绕 {highlights[0]['product_name']} 的当日优势制作简洁的事实卡片。")
        if len(news) >= 2:
            suggestions.append("将当日重点信号拆分为“市场变化 + 对应解读”两段式内容。")
        suggestions.append("对外内容保留事实依据与风险提示，避免夸张表述。")

        return {
            "available": True,
            "date": date_str,
            "highlights": highlights,
            "news": news,
            "suggestions": suggestions[:4],
        }

    def get_bootstrap_state(self, simulation_state: dict) -> dict:
        state = self.get_current_state()
        current_date = state["date"] if state else None
        workspace = self.profile.get("workspace", {})
        modules = self._module_definitions()
        competition_story = self.get_competition_story()
        return {
            "app_name": workspace.get("app_name", "资管产品洞察协作台"),
            "workspace": workspace,
            "modules": [{**item, "skills": self.module_skill_cards(item["key"])} for item in modules],
            "has_data": self.has_data(),
            "mode": self.mode,
            "warnings": self.warnings,
            "simulation": {
                "is_running": simulation_state["is_running"],
                "interval": simulation_state["interval"],
                "current_date": current_date,
            },
            "summary": {
                "product_count": len(self.get_all_products()),
                "policy_count": len(self.policies),
                "signal_count": len(self.all_signals.get(current_date, [])) if current_date else 0,
                "status": "演示模式" if self.mode == "demo" else ("已接入" if self.has_data() else "空态"),
            },
            "current_state": state
            or {
                "date": None,
                "products": [],
                "indices": {},
                "policies": [],
            },
            "products": self.get_all_products(),
            "signals": self.all_signals.get(current_date, []) if current_date else [],
            "history": {date: self.all_signals.get(date, []) for date in reversed(self.dates[-10:])},
            "daily_report": self.get_daily_report(current_date),
            "demo_cases": self.get_demo_cases(),
            "trace_summary": self.default_trace_summary(),
            "quality_metrics": self.default_quality_metrics(),
            "comparison": self.default_comparison(),
            "competition_story": competition_story,
            "research_overview": self.get_research_overview(),
            "impact_path": self.get_impact_path(),
            "diagnostics": self.get_diagnostics(),
            "validation_report": self.get_validation_report(),
        }

    def get_competition_story(self) -> dict:
        competition = self.profile.get("competition", {})
        stages = competition.get("stages") or [
            {"title": "数据接入", "summary": "加载市场快照、政策目录和文档材料。"},
            {"title": "多源分析", "summary": "结合规则信号、RAG 和专家 skills 做归因。"},
            {"title": "结构化结论", "summary": "把事实、判断、建议和风险分层输出。"},
            {"title": "报告输出", "summary": "生成研判报告、质量评分和答辩大纲。"},
        ]
        quant_outputs_raw = competition.get("quant_outputs") or [
            {"label": "岗位化模块", "value": str(len(MODULES))},
            {"label": "专家 Skills", "value": str(len(self.module_skill_cards("policy_analysis")) + 3)},
            {"label": "输出形态", "value": "3"},
        ]
        quant_outputs = []
        for item in quant_outputs_raw:
            if isinstance(item, dict):
                quant_outputs.append({"label": str(item.get("label", "")), "value": str(item.get("value", ""))})
            elif isinstance(item, list) and len(item) >= 2:
                quant_outputs.append({"label": str(item[0]), "value": str(item[1])})
        return {
            "headline": competition.get("headline", "面向资管研究与政策研判的智能分析工作台"),
            "subheadline": competition.get(
                "subheadline",
                "把多源数据研究、专家技能编排和结构化研报输出压缩进同一条演示闭环。",
            ),
            "stages": stages,
            "why_it_matters": competition.get("why_it_matters", []),
            "differentiators": competition.get("differentiators", []),
            "demo_flow": competition.get("demo_flow", []),
            "quant_outputs": quant_outputs,
        }

    def get_demo_cases(self) -> List[dict]:
        templates = self.profile.get("competition", {}).get("demo_cases", [])
        cases: List[dict] = []
        current_date = self.get_current_date() or (self.dates[0] if self.dates else None)
        state = self.get_current_state() or {"products": [], "indices": {}, "policies": []}
        signals = self.all_signals.get(current_date, []) if current_date else []
        top_product = self.get_market_leaderboard("inflow", top_n=1)
        lead_product = top_product[0] if top_product else None
        top_signal = signals[0] if signals else None
        top_policy = (state.get("policies") or self.policies[:1] or [None])[0]

        for template in templates:
            case_id = template.get("id", "demo_case")
            active_module = template.get("active_module", "policy_analysis")
            case = {
                "id": case_id,
                "title": template.get("title", case_id),
                "category": template.get("category", "演示案例"),
                "hook": template.get("hook", "展示多源分析到结构化报告的完整闭环。"),
                "summary": template.get("summary", "点击加载比赛演示案例。"),
                "tags": template.get("tags", []),
                "active_module": active_module,
                "focus_date": current_date,
                "available": bool(current_date),
                "button_label": template.get("button_label", "加载案例"),
                "key_question": template.get("key_question", "请给出多源研判和结构化输出。"),
                "input_data": [],
            }

            if case_id == "policy_shock":
                policy_title = top_policy["标题"] if top_policy else "示例监管通知"
                case["summary"] = template.get("summary", f"围绕政策条目“{policy_title}”展示政策研判闭环。")
                case["key_question"] = template.get(
                    "key_question",
                    f"请围绕政策条目“{policy_title}”，说明关键变化、市场传导和产品动作。",
                )
                case["input_data"] = [
                    f"政策目录：{policy_title}",
                    f"研究日期：{current_date or '未加载'}",
                    f"可用信号：{len(signals)} 条",
                ]
            elif case_id == "market_volatility":
                signal_content = top_signal["content"] if top_signal else "示例指数出现波动"
                case["summary"] = template.get("summary", f"围绕市场异动“{signal_content}”展示监测到策略输出。")
                case["key_question"] = template.get(
                    "key_question",
                    f"请解释“{signal_content}”的市场影响，并形成研究摘要与行动建议。",
                )
                case["input_data"] = [
                    f"当日信号：{signal_content}",
                    f"历史回放：{len(self.dates)} 个日期",
                    f"可用产品：{len(state.get('products', []))} 个",
                ]
            elif case_id == "product_strategy":
                product_name = lead_product["name"] if lead_product else "示例产品"
                case["summary"] = template.get("summary", f"围绕重点产品“{product_name}”展示研究到策略的完整闭环。")
                case["key_question"] = template.get(
                    "key_question",
                    f"请围绕重点产品“{product_name}”，给出研究摘要、市场判断和策略动作。",
                )
                case["input_data"] = [
                    f"重点产品：{product_name}",
                    f"重点日期：{current_date or '未加载'}",
                    f"当日政策：{len(state.get('policies', []))} 条",
                ]
            else:
                case["input_data"] = [f"研究日期：{current_date or '未加载'}"]

            cases.append(case)

        return cases

    def get_research_overview(self) -> dict:
        state = self.get_current_state()
        if not state:
            return {
                "headline": "等待研究对象",
                "items": [
                    {"label": "模式", "value": "空态模式"},
                    {"label": "记录", "value": "0"},
                    {"label": "政策", "value": "0"},
                    {"label": "信号", "value": "0"},
                ],
            }

        products = state.get("products", [])
        policies = state.get("policies", [])
        signals = self.all_signals.get(state["date"], [])
        leader = max(products, key=lambda item: item.get("inflow", 0), default=None)
        return {
            "headline": f"{state['date']} 研究对象概览",
            "items": [
                {"label": "运行模式", "value": "演示模式" if self.mode == "demo" else "外部数据"},
                {"label": "重点产品", "value": leader["name"] if leader else "暂无"},
                {"label": "政策条目", "value": str(len(policies))},
                {"label": "当日信号", "value": str(len(signals))},
            ],
        }

    def get_impact_path(self) -> dict:
        state = self.get_current_state()
        signals = self.all_signals.get(self.get_current_date() or "", [])
        top_policy = (state or {}).get("policies", [])[:1]
        top_signal = signals[:1]
        steps = [
            {
                "title": "输入数据",
                "detail": top_policy[0]["标题"] if top_policy else (top_signal[0]["content"] if top_signal else "市场快照 / 政策目录 / 文档材料"),
            },
            {
                "title": "多源分析",
                "detail": "规则信号、RAG 检索与专家 skills 协同完成事实归因。",
            },
            {
                "title": "结构化结论",
                "detail": "区分事实、判断、建议和风险边界，避免直接生成式漂移。",
            },
            {
                "title": "结果输出",
                "detail": "同步生成研究摘要、正式报告和答辩大纲。",
            },
        ]
        return {"title": "政策 / 信号影响路径", "steps": steps}

    def get_diagnostics(self) -> dict:
        checks = [
            {
                "label": "语言模型",
                "status": "ready" if self._get_llm_api_key() else "fallback",
                "detail": "已接入模型服务" if self._get_llm_api_key() else "未配置，当前将使用离线摘要与报告模板",
            },
            {
                "label": "Embedding",
                "status": "ready" if self._embedding_available() else "fallback",
                "detail": "支持语义检索与意图识别" if self._embedding_available() else "未配置，当前退回关键词与规则检索",
            },
            {
                "label": "数据模式",
                "status": "ready" if self.has_data() else "waiting",
                "detail": {
                    "external": "当前加载外部数据目录",
                    "demo": "当前加载内置演示数据",
                }.get(self.mode, "当前未加载业务数据"),
            },
            {
                "label": "Profile",
                "status": "ready",
                "detail": f"当前配置：{self._profile_path() or (self.base_dir / 'config' / 'default_profile.json')}",
            },
        ]
        return {"mode": self.mode, "checks": checks}

    def get_validation_report(self) -> dict:
        products = [item for date in self.dates for item in self.data.get(date, {}).get("products", [])]
        indices = [item for date in self.dates for item in self.data.get(date, {}).get("indices", {}).values()]
        missing_setup = sum(1 for item in products if not item.get("setup_date"))
        missing_index = sum(1 for item in products if not item.get("index_code"))
        return {
            "mode": self.mode,
            "date_count": len(self.dates),
            "record_count": len(products),
            "index_count": len(indices),
            "policy_count": len(self.policies),
            "missing_setup_ratio": round(missing_setup / len(products), 2) if products else 0.0,
            "missing_index_ratio": round(missing_index / len(products), 2) if products else 0.0,
            "usable": bool(self.has_data() or self.policies),
        }

    def default_trace_summary(self) -> dict:
        skills = get_skill_cards(self._workflow_skill_keys("report"))
        return {
            "title": "增强式研报编排",
            "subtitle": "加载 Demo Case 或上传文件后，将展示专家 skills、证据来源和纳入报告的关键观察。",
            "steps": [
                {"title": "数据接入", "detail": "加载市场快照、政策目录和文档内容。"},
                {"title": "技能拆解", "detail": "按政策解读、市场影响、产品策略、风险合规、报告编审拆解。"},
                {"title": "证据归并", "detail": "把市场快照、政策目录、历史信号和文档内容统一归因。"},
                {"title": "报告成稿", "detail": "输出带来源标签的结构化研判报告。"},
            ],
            "skills": [
                {"label": item["label"], "focus": item["focus"], "observations": []}
                for item in skills
            ],
            "evidence": [
                {"category": "市场快照", "highlights": []},
                {"category": "政策目录", "highlights": []},
                {"category": "历史信号", "highlights": []},
                {"category": "上传文档", "highlights": []},
            ],
            "included_observations": [],
            "report_sections": ["政策概要", "关键变化", "市场与业务影响", "产品策略观察", "风险与合规边界", "执行优先级"],
        }

    def default_quality_metrics(self) -> List[dict]:
        readiness_bonus = 20 if self.has_data() else 0
        llm_bonus = 20 if self._get_llm_api_key() else 0
        embed_bonus = 15 if self._embedding_available() else 0
        metrics = [
            ("fact_coverage", "事实覆盖度", 45 + readiness_bonus, "结构化数据和证据基础已就绪"),
            ("citation_rate", "证据引用率", 40 + readiness_bonus + embed_bonus, "支持市场快照、政策目录、历史信号和文档来源标记"),
            ("risk_completeness", "风险提示完整度", 55 + llm_bonus // 2, "风险与合规 skill 已加入默认工作流"),
            ("structure_integrity", "结构完整度", 70 + llm_bonus // 2, "输出固定为摘要、报告和答辩大纲三类结果"),
        ]
        return [
            {
                "key": key,
                "label": label,
                "score": min(100, score),
                "baseline_score": max(20, score - 18),
                "status": self._quality_status(score),
                "detail": detail,
            }
            for key, label, score, detail in metrics
        ]

    def default_comparison(self) -> dict:
        baseline = {
            "label": "Baseline",
            "summary": "直接生成结果，结构较短，缺少来源标签和技能拆解。",
            "sections": 4,
            "citations": 0,
            "score": 58,
        }
        enhanced = {
            "label": "Enhanced",
            "summary": "结合规则信号、RAG 和专家 skills，输出可解释的结构化结论。",
            "sections": 6,
            "citations": 4,
            "score": 84,
        }
        return {
            "title": "Baseline vs Enhanced",
            "baseline": baseline,
            "enhanced": enhanced,
            "diffs": [
                "输出结构从 4 段提升到 6 段。",
                "证据来源从 0 提升到 4 类。",
                "额外补充风险边界与执行优先级。",
            ],
        }

    def build_demo_case_artifact(self, case_id: str) -> dict:
        case = next((item for item in self.get_demo_cases() if item["id"] == case_id), None)
        if not case or not case.get("available"):
            raise ValueError("当前案例不可用，请先接入数据或启用演示模式。")

        if case.get("focus_date") in self.dates:
            self.current_index = self.dates.index(case["focus_date"])

        briefing_text = self._compose_demo_case_briefing(case)
        artifact = self.build_document_artifact(briefing_text, case=case)
        artifact["case"] = case
        artifact["assistant_message"] = self._build_demo_case_message(case, artifact)
        return artifact

    def build_document_artifact(self, text: str, case: Optional[dict] = None) -> dict:
        trace = self._build_trace_summary(text, case=case)
        baseline_report = self._build_baseline_report(text, trace, case=case)
        raw_enhanced = self.analyze_document(text)
        enhanced_report = self._compose_enhanced_report(raw_enhanced, trace, case=case)
        baseline_summary = self._report_to_summary(baseline_report)
        enhanced_summary = self._report_to_summary(enhanced_report)
        quality_metrics = self._build_quality_metrics(baseline_report, enhanced_report, trace)
        comparison = self._build_comparison_snapshot(baseline_summary, enhanced_summary, baseline_report, enhanced_report, quality_metrics)
        outline = self._build_competition_outline(case, enhanced_summary, quality_metrics, trace)
        return {
            "text": text,
            "baseline_report": baseline_report,
            "enhanced_report": enhanced_report,
            "baseline_summary": baseline_summary,
            "enhanced_summary": enhanced_summary,
            "trace_summary": trace,
            "quality_metrics": quality_metrics,
            "comparison": comparison,
            "outline": outline,
            "report_title": case.get("report_title", "政策研判报告") if case else "政策研判报告",
        }

    def _build_demo_case_message(self, case: dict, artifact: dict) -> str:
        return (
            f"### {case['title']}\n\n"
            f"- 案例类型：{case['category']}\n"
            f"- 关键问题：{case['key_question']}\n"
            f"- 输入数据：{'；'.join(case.get('input_data', []))}\n"
            f"- 结构化结论：{artifact['enhanced_summary'].splitlines()[0] if artifact['enhanced_summary'] else '已生成增强版摘要'}\n"
            f"- 输出结果：已生成正式报告、质量评分卡与答辩大纲。"
        )

    def _compose_demo_case_briefing(self, case: dict) -> str:
        state = self.get_current_state() or {"date": None, "products": [], "indices": {}, "policies": []}
        signals = self.all_signals.get(state.get("date") or "", [])
        products = state.get("products", [])
        indices = list(state.get("indices", {}).values())
        top_product = max(products, key=lambda item: item.get("inflow", 0), default=None)
        top_signal = signals[0] if signals else None
        top_policy = (state.get("policies") or [None])[0]
        lines = [
            f"案例名称：{case['title']}",
            f"案例类型：{case['category']}",
            f"研究日期：{state.get('date') or case.get('focus_date') or '未加载'}",
            f"关键问题：{case['key_question']}",
            "",
            "一、输入数据",
        ]
        lines.extend(f"- {item}" for item in case.get("input_data", []))
        lines.extend(["", "二、多源事实"])
        if top_product:
            lines.append(
                f"- 市场快照显示，{top_product['name']}({top_product['code']}) 当日净流入 {top_product['inflow']:.2f} 亿，成交额 {top_product['volume']:.2f} 亿。"
            )
        if indices:
            lead_index = max(indices, key=lambda item: abs(item.get("change", 0)))
            lines.append(f"- 指数侧最显著变化来自 {lead_index['name']}，日内涨跌幅 {lead_index.get('change', 0):.1f}%。")
        if top_policy:
            lines.append(f"- 政策目录中可关联条目为《{top_policy['标题']}》，位阶为 {top_policy['法律位阶']}。")
        if top_signal:
            lines.append(f"- 历史信号提示：{top_signal['content']}")
        lines.extend(
            [
                "",
                "三、任务要求",
                "- 先做事实归纳，再说明影响路径，最后给出产品动作与风险边界。",
                "- 输出需支持正式研判报告和答辩摘要两个结果物。",
            ]
        )
        return "\n".join(lines)

    def _build_trace_summary(self, text: str, case: Optional[dict] = None) -> dict:
        evidence = self._collect_evidence(text, case=case)
        skill_cards = get_skill_cards(self._workflow_skill_keys("report"))
        skills = []
        included_observations: List[str] = []
        for card in skill_cards:
            observations = self._skill_observations(card["key"], evidence, case=case)
            included_observations.extend(observations[:1])
            skills.append(
                {
                    "label": card["label"],
                    "focus": card["focus"],
                    "observations": observations,
                }
            )
        return {
            "title": case["title"] if case else "增强式研报编排",
            "subtitle": case["key_question"] if case else "展示专家协作、证据来源与结构化成稿路径。",
            "steps": [
                {"title": "数据接入", "detail": "整合市场快照、政策目录与文档上下文。"},
                {"title": "专家拆解", "detail": "按政策解读、市场影响、产品策略、风险合规与报告编审拆解。"},
                {"title": "证据归并", "detail": "将多类证据统一映射到报告章节。"},
                {"title": "结构化成稿", "detail": "输出带来源标签的六段式研判报告。"},
            ],
            "skills": skills,
            "evidence": evidence,
            "included_observations": included_observations[:6],
            "report_sections": ["政策概要", "关键变化", "市场与业务影响", "产品策略观察", "风险与合规边界", "执行优先级"],
        }

    def _collect_evidence(self, text: str, case: Optional[dict] = None) -> List[dict]:
        state = self.get_current_state() or {"products": [], "indices": {}, "policies": [], "date": None}
        signals = self.all_signals.get(state.get("date") or "", [])
        products = state.get("products", [])
        indices = list(state.get("indices", {}).values())
        product_highlights = [
            f"{item['name']} 净流入 {item['inflow']:.2f} 亿 / 成交额 {item['volume']:.2f} 亿"
            for item in sorted(products, key=lambda entry: entry.get("inflow", 0), reverse=True)[:2]
        ]
        index_highlights = [
            f"{item['name']} 涨跌幅 {item.get('change', 0):.1f}%"
            for item in sorted(indices, key=lambda entry: abs(entry.get("change", 0)), reverse=True)[:1]
        ]
        policy_highlights = [
            f"{item['标题']} ({item['法律位阶']})"
            for item in (state.get("policies") or self.policies[:2])[:2]
            if item.get("标题")
        ]
        signal_highlights = [item["content"] for item in signals[:3]]
        doc_highlights = self._extract_document_clauses(text, limit=3)
        return [
            {
                "category": "市场快照",
                "highlights": product_highlights + index_highlights,
            },
            {
                "category": "政策目录",
                "highlights": policy_highlights,
            },
            {
                "category": "历史信号",
                "highlights": signal_highlights,
            },
            {
                "category": "上传文档" if case else "文档内容",
                "highlights": doc_highlights,
            },
        ]

    def _skill_observations(self, skill_key: str, evidence: List[dict], case: Optional[dict] = None) -> List[str]:
        evidence_map = {item["category"]: item["highlights"] for item in evidence}
        market_points = evidence_map.get("市场快照", [])
        policy_points = evidence_map.get("政策目录", [])
        history_points = evidence_map.get("历史信号", [])
        document_points = evidence_map.get("上传文档", []) or evidence_map.get("文档内容", [])
        if skill_key == "policy_interpreter":
            return [
                document_points[0] if document_points else "政策文本可提炼出核心约束条件和执行对象。",
                policy_points[0] if policy_points else "政策目录为当前研判提供了条目级背景。",
            ]
        if skill_key == "market_impact_analyst":
            return [
                history_points[0] if history_points else "历史信号提示市场已有明确反应。",
                market_points[-1] if market_points else "市场快照可用于观察指数或板块传导路径。",
            ]
        if skill_key == "product_strategy_advisor":
            return [
                market_points[0] if market_points else "重点产品表现可以作为策略动作依据。",
                "建议把产品动作限定在可量化、可验证的业务抓手上。",
            ]
        if skill_key == "risk_compliance_reviewer":
            return [
                "所有判断均需保留适用范围、时间窗口和信息完整性的前提说明。",
                "对外表述应避免确定性收益和过度推演。",
            ]
        if skill_key == "report_editor":
            return [
                "报告统一拆为政策概要、关键变化、市场影响、策略观察、风险边界、执行优先级六段。",
                "各章节均附来源类别，方便答辩时展示可解释性。",
            ]
        return [
            market_points[0] if market_points else "多源数据已进入工作流。",
            document_points[0] if document_points else "文档信息可作为补充证据。",
        ]

    def _build_baseline_report(self, text: str, trace: dict, case: Optional[dict] = None) -> str:
        clauses = self._extract_document_clauses(text, limit=4)
        while len(clauses) < 4:
            clauses.append("当前信息有限，建议结合更多数据与正式文本继续复核。")
        title = case["title"] if case else "直接生成草稿"
        return (
            f"# {title}\n"
            f"{clauses[0]}\n\n"
            "# 主要判断\n"
            f"{clauses[1]}\n\n"
            "# 业务提示\n"
            f"{clauses[2]}\n\n"
            "# 风险提醒\n"
            f"{clauses[3]}"
        )

    def _compose_enhanced_report(self, raw_report: str, trace: dict, case: Optional[dict] = None) -> str:
        source_map = {
            "政策概要": ["上传文档", "政策目录"],
            "关键变化": ["上传文档", "政策目录"],
            "市场与业务影响": ["市场快照", "历史信号"],
            "产品策略观察": ["市场快照", "历史信号"],
            "风险与合规边界": ["上传文档", "政策目录"],
            "执行优先级": ["市场快照", "历史信号", "上传文档"],
        }
        extracted = self._extract_report_sections(raw_report)
        fallback = self._fallback_report_sections(trace, case=case)
        sections = []
        for heading in trace["report_sections"]:
            body = extracted.get(heading) or fallback.get(heading) or "当前信息有限，建议结合更多证据继续补充。"
            sources = " / ".join(source_map.get(heading, ["上传文档"]))
            sections.append(f"# {heading}\n{body}\n[来源: {sources}]")
        return "\n\n".join(sections)

    def _extract_report_sections(self, raw_report: str) -> Dict[str, str]:
        sections: Dict[str, List[str]] = {}
        current_heading: Optional[str] = None
        for line in raw_report.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            if stripped.startswith("#"):
                current_heading = stripped.replace("#", "").strip()
                sections.setdefault(current_heading, [])
                continue
            if current_heading:
                sections[current_heading].append(stripped)
        return {key: " ".join(value).strip() for key, value in sections.items()}

    def _fallback_report_sections(self, trace: dict, case: Optional[dict] = None) -> Dict[str, str]:
        skill_map = {item["label"]: item["observations"] for item in trace.get("skills", [])}
        policy_obs = skill_map.get("政策解读", [])
        market_obs = skill_map.get("宏观与市场影响", [])
        strategy_obs = skill_map.get("产品策略建议", [])
        risk_obs = skill_map.get("风险与合规", [])
        return {
            "政策概要": policy_obs[0] if policy_obs else "本次案例聚焦政策或事件的核心变化及其研究价值。",
            "关键变化": policy_obs[1] if len(policy_obs) > 1 else "关键变化体现在约束条件、适用对象或执行要求上。",
            "市场与业务影响": "；".join(market_obs[:2]) if market_obs else "市场影响主要通过流动性、风险偏好和交易节奏传导。",
            "产品策略观察": "；".join(strategy_obs[:2]) if strategy_obs else "产品策略建议应聚焦对象、动作和触发条件。",
            "风险与合规边界": "；".join(risk_obs[:2]) if risk_obs else "需明确适用边界、数据完整性和表述限制。",
            "执行优先级": "优先完成事实核对、影响归因、策略筛选和风险复核，再形成正式输出。",
        }

    def _report_to_summary(self, report: str) -> str:
        extracted = self._extract_report_sections(report)
        headings = ["政策概要", "关键变化", "市场与业务影响", "风险与合规边界"]
        lines = []
        for idx, heading in enumerate(headings, 1):
            body = extracted.get(heading, "当前暂无对应内容。")
            lines.append(f"{idx}. **{heading}**：{body[:90]}")
        return "\n".join(lines)

    def _build_quality_metrics(self, baseline_report: str, enhanced_report: str, trace: dict) -> List[dict]:
        baseline_sections = max(1, baseline_report.count("# "))
        enhanced_sections = max(1, enhanced_report.count("# "))
        citations = enhanced_report.count("[来源:")
        evidence_highlights = sum(len(item.get("highlights", [])) for item in trace.get("evidence", []))
        risk_points = enhanced_report.count("风险") + enhanced_report.count("边界")
        metrics = [
            {
                "key": "fact_coverage",
                "label": "事实覆盖度",
                "baseline_score": min(100, 36 + baseline_sections * 8),
                "score": min(100, 52 + min(24, evidence_highlights * 5) + enhanced_sections * 4),
                "detail": "综合结构化数据、政策目录、历史信号与文档内容的覆盖情况。",
            },
            {
                "key": "citation_rate",
                "label": "证据引用率",
                "baseline_score": 18,
                "score": min(100, 34 + citations * 14),
                "detail": "报告章节是否显式标注来源类别。",
            },
            {
                "key": "risk_completeness",
                "label": "风险提示完整度",
                "baseline_score": 32,
                "score": min(100, 48 + risk_points * 8),
                "detail": "是否补充适用边界、信息缺口和过度解读风险。",
            },
            {
                "key": "structure_integrity",
                "label": "结构完整度",
                "baseline_score": min(100, 25 + baseline_sections * 10),
                "score": min(100, 28 + enhanced_sections * 11),
                "detail": "是否完整覆盖摘要、影响、策略、风险和执行优先级。",
            },
        ]
        for metric in metrics:
            metric["status"] = self._quality_status(metric["score"])
        return metrics

    def _build_comparison_snapshot(
        self,
        baseline_summary: str,
        enhanced_summary: str,
        baseline_report: str,
        enhanced_report: str,
        quality_metrics: List[dict],
    ) -> dict:
        overall_baseline = round(sum(item["baseline_score"] for item in quality_metrics) / len(quality_metrics))
        overall_enhanced = round(sum(item["score"] for item in quality_metrics) / len(quality_metrics))
        return {
            "title": "Baseline vs Enhanced",
            "baseline": {
                "label": "Baseline",
                "summary": baseline_summary.splitlines()[0] if baseline_summary else "直接生成的简版草稿。",
                "sections": baseline_report.count("# "),
                "citations": baseline_report.count("[来源:"),
                "score": overall_baseline,
            },
            "enhanced": {
                "label": "Enhanced",
                "summary": enhanced_summary.splitlines()[0] if enhanced_summary else "专家 skills 编排增强版本。",
                "sections": enhanced_report.count("# "),
                "citations": enhanced_report.count("[来源:"),
                "score": overall_enhanced,
            },
            "diffs": [
                f"整体评分提升 {overall_enhanced - overall_baseline} 分。",
                f"章节数量从 {baseline_report.count('# ')} 提升到 {enhanced_report.count('# ')}。",
                f"来源标记从 {baseline_report.count('[来源:')} 提升到 {enhanced_report.count('[来源:')}。",
            ],
        }

    def _build_competition_outline(self, case: Optional[dict], summary: str, quality_metrics: List[dict], trace: dict) -> str:
        story = self.get_competition_story()
        overall_score = round(sum(item["score"] for item in quality_metrics) / len(quality_metrics)) if quality_metrics else 0
        lines = [
            f"# {self.profile.get('workspace', {}).get('app_name', '资管产品洞察协作台')} 答辩摘要",
            "",
            "## 项目定位",
            story["headline"],
            story["subheadline"],
            "",
            "## Demo 场景",
            case["title"] if case else "上传文档 / 即时研判",
            case["key_question"] if case else "围绕当前输入完成多源分析和结构化报告输出。",
            "",
            "## AI 技术方案",
            "- 结构化数据处理 + RAG 检索增强 + 金融专家 skills 编排 + 结构化报告导出。",
            "- 使用市场快照、政策目录、历史信号和文档内容四类证据完成归因。",
            "",
            "## 关键结论",
            summary,
            "",
            "## 可解释性证据",
            *[f"- {item}" for item in trace.get("included_observations", [])[:4]],
            "",
            "## 质量评分",
            *[f"- {item['label']}：{item['score']} 分" for item in quality_metrics],
            f"- 综合评分：{overall_score} 分",
            "",
            "## 竞赛价值",
            *[f"- {item}" for item in story.get("why_it_matters", [])[:3]],
        ]
        return "\n".join(lines)

    def build_outline_bytes(self, content: str) -> bytes:
        return content.encode("utf-8")

    def _quality_status(self, score: int) -> str:
        thresholds = self.profile.get("competition", {}).get("quality_thresholds", {})
        strong = int(thresholds.get("strong", 85))
        watch = int(thresholds.get("watch", 65))
        if score >= strong:
            return "strong"
        if score >= watch:
            return "watch"
        return "weak"

    def _embedding_available(self) -> bool:
        return bool(os.getenv("EMBEDDING_API_KEY", "").strip() or os.getenv("DASHSCOPE_API_KEY", "").strip())

    def _module_definitions(self) -> List[dict]:
        overrides = self.profile.get("workspace", {}).get("module_overrides", {})
        definitions = []
        for item in MODULES:
            override = overrides.get(item["key"], {}) if isinstance(overrides, dict) else {}
            merged = dict(item)
            if isinstance(override, dict):
                merged.update({key: value for key, value in override.items() if value})
            definitions.append(merged)
        return definitions

    def _module_map(self) -> Dict[str, dict]:
        return {item["key"]: item for item in self._module_definitions()}

    def find_product(self, query: str, date_str: Optional[str] = None) -> Optional[dict]:
        target_date = date_str or self.get_current_date()
        if not target_date or target_date not in self.data:
            return None

        query_lower = query.lower().strip()
        products = self.data[target_date].get("products", [])
        for item in products:
            if item["code"].lower() == query_lower:
                return item
        for item in products:
            if query_lower in item["name"].lower():
                return item
        return None

    def render_product_snapshot(self, product: dict, date_str: str) -> str:
        return (
            f"### 产品快照\n\n"
            f"- 日期：{date_str}\n"
            f"- 名称：{product['name']}\n"
            f"- 代码：{product['code']}\n"
            f"- 规模：{product['scale']:.2f} 亿\n"
            f"- 成交额：{product['volume']:.2f} 亿\n"
            f"- 净流入：{product['inflow']:.2f} 亿"
        )

    def render_leaderboard(self, category: str) -> str:
        label_map = {"volume": ("成交活跃", "成交额"), "inflow": ("资金流入", "净流入"), "scale": ("规模领先", "规模")}
        category_label, metric_label = label_map[category]
        products = self.get_market_leaderboard(category, top_n=3)
        if not products:
            return "当前无可用排行数据。"

        date_str = self.get_current_date() or "-"
        rows = "\n".join(
            f"| {idx} | {item['name']} ({item['code']}) | {item[category]:.2f} 亿 |"
            for idx, item in enumerate(products, 1)
        )
        return (
            f"### {date_str} {category_label}排行榜\n\n"
            f"| 排名 | 产品名称 | {metric_label} |\n"
            f"|:---|:---|---:|\n"
            f"{rows}"
        )

    def detect_structured_response(self, message: str, module_key: str) -> Optional[str]:
        if not self.has_data():
            if module_key == "policy_analysis":
                return None
            return "当前未加载外部数据目录。请先配置 `DATA_SOURCE_DIR`，或在空态下使用政策解析能力。"

        code_match = re.search(r"([0-9]{6}(?:\.[A-Z]{2})?)", message)
        if code_match:
            product = self.find_product(code_match.group(1))
            if product:
                return self.render_product_snapshot(product, self.get_current_date() or "-")

        if any(keyword in message for keyword in ["成交活跃", "成交额排行", "谁成交最多", "成交额排名"]):
            return self.render_leaderboard("volume")

        if any(keyword in message for keyword in ["净流入", "资金流入", "谁吸金", "流入排行"]):
            return self.render_leaderboard("inflow")

        if any(keyword in message for keyword in ["规模领先", "规模排行", "规模排名"]):
            return self.render_leaderboard("scale")

        range_match = re.search(r"(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})\s*[到至\-~]\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})", message)
        if range_match:
            return self.get_date_range_comparison(
                self._normalize_date(range_match.group(1)) or range_match.group(1),
                self._normalize_date(range_match.group(2)) or range_match.group(2),
                code_match.group(1) if code_match else None,
            )

        date_match = re.search(r"(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})", message)
        if date_match:
            normalized_date = self._normalize_date(date_match.group(1))
            if normalized_date:
                return self.get_date_summary(normalized_date)

        return None

    def module_label(self, module_key: str) -> str:
        return self._module_map().get(module_key, MODULES[0])["label"]

    def module_skill_cards(self, module_key: str) -> List[dict]:
        return get_module_skill_cards(module_key)

    def _workflow_skill_keys(self, workflow_key: str) -> List[str]:
        return DOCUMENT_WORKFLOWS.get(workflow_key, [])

    def build_system_prompt(self, module_key: str) -> str:
        module_info = self._module_map().get(module_key, MODULES[0])
        context = self.get_daily_summary()
        skillbook = build_skillbook(MODULE_SKILLS.get(module_key, []))
        return (
            f"{module_info.get('prompt', MODULES[0]['prompt'])}\n\n"
            f"当前启用的专家 skills：\n{skillbook}\n\n"
            f"回答要求：综合这些技能视角作答，但不要机械罗列技能名；明确区分事实、判断、建议和风险边界。\n\n"
            f"当前上下文：\n{context}"
        )

    def build_llm_prompt(self, message: str, module_key: str) -> str:
        intent = self.get_user_intent(message)
        history_context = ""
        if intent in {"planning", "comparison", "policy"}:
            similar = self.get_similar_signals(message, top_k=2)
            if similar:
                history_context = "\n\n历史参考：\n" + "\n".join(
                    f"- [{item['date']}] {item['content']}" for item in similar
                )
        state_date = self.get_current_date() or "未加载数据"
        skillbook = build_skillbook(MODULE_SKILLS.get(module_key, []), include_contract=False)
        return (
            f"当前日期：{state_date}\n"
            f"模块：{self.module_label(module_key)}\n"
            f"当前技能编排：\n{skillbook}\n\n"
            f"用户问题：{message}{history_context}"
        )

    def check_rules(self) -> List[dict]:
        state = self.get_current_state()
        if not state:
            return []

        date_str = state["date"]
        products = state["products"]
        indices = state["indices"]
        signals: List[dict] = []

        ranked_scale = sorted(products, key=lambda item: item.get("scale", 0), reverse=True)
        for idx, item in enumerate(ranked_scale[:3], 1):
            signals.append(
                {
                    "type": "scale",
                    "category": "规模里程碑",
                    "content": f"{item['name']} 规模位列样本产品第 {idx}，当前为 {item['scale']:.2f} 亿。",
                    "date": date_str,
                }
            )

        for index_item in indices.values():
            change = index_item.get("change", 0)
            if abs(change) >= 1.5:
                direction = "上涨" if change > 0 else "下跌"
                signals.append(
                    {
                        "type": "market",
                        "category": "行情异动",
                        "content": f"{index_item['name']} 日内{direction} {abs(change):.1f}%。",
                        "date": date_str,
                    }
                )

        top_inflow = sorted(products, key=lambda item: item.get("inflow", 0), reverse=True)[:3]
        for item in top_inflow:
            if item.get("inflow", 0) > 0:
                signals.append(
                    {
                        "type": "product",
                        "category": "产品动态",
                        "content": f"{item['name']} 当日净流入 {item['inflow']:.2f} 亿。",
                        "date": date_str,
                    }
                )

        current_dt = datetime.strptime(date_str, "%Y-%m-%d")
        for item in products:
            setup_date = item.get("setup_date")
            if not setup_date:
                continue
            try:
                setup_dt = datetime.strptime(setup_date, "%Y-%m-%d")
            except ValueError:
                continue
            if setup_dt.month == current_dt.month and setup_dt.day == current_dt.day:
                years = current_dt.year - setup_dt.year
                if years > 0:
                    signals.append(
                        {
                            "type": "anniversary",
                            "category": "周年提醒",
                            "content": f"{item['name']} 迎来成立 {years} 周年。",
                            "date": date_str,
                        }
                    )

        for item in self.policies:
            if item.get("公告日期") == date_str and item.get("标题"):
                signals.append(
                    {
                        "type": "policy",
                        "category": "监管政策",
                        "content": f"[{item['法律位阶']}] {item['标题']}",
                        "date": date_str,
                        "link": item.get("来源", ""),
                    }
                )

        return signals

    def summarize_document(self, text: str) -> str:
        if not self._get_llm_api_key():
            return self._offline_document_summary(text)

        skill_notes = self._generate_document_skill_notes(text, workflow_key="summary")
        prompt = (
            "请基于以下金融专家 skills 分析笔记，为业务用户输出一份四点式政策解析摘要。"
            "每一点都要同时覆盖条款变化、影响路径或注意事项，使用“标题：说明”的形式，"
            "说明控制在 60 至 90 字。\n\n"
            f"技能笔记：\n{skill_notes}\n\n"
            f"原文节选：\n{text[:3200]}"
        )
        return self.call_llm(
            prompt,
            system_content=(
                "你是政策解析模块的编审器。你的工作是综合金融专家 skills 笔记，"
                "输出短而准的摘要，避免空话和确定性过强的表述。"
            ),
            model="qwen-flash",
        )

    def analyze_document(self, text: str) -> str:
        if not self._get_llm_api_key():
            return self._offline_document_report(text)

        skill_notes = self._generate_document_skill_notes(text, workflow_key="report")
        prompt = (
            "请基于以下金融专家 skills 笔记和文件原文，生成一份结构化研判报告。"
            "报告必须包含以下层级：\n"
            "# 政策概要\n# 关键变化\n# 市场与业务影响\n# 产品策略观察\n# 风险与合规边界\n# 执行优先级\n\n"
            "写作要求：\n"
            "1. 先事实，后判断，再建议。\n"
            "2. 避免宏大口号和夸张收益表述。\n"
            "3. 每个章节都要给出明确落点。\n\n"
            f"技能笔记：\n{skill_notes}\n\n"
            f"原文节选：\n{text[:5000]}"
        )
        return self.call_llm(
            prompt,
            system_content=(
                "你是政策解析模块的报告编审器，负责融合多位金融专家 skill 的观点，"
                "输出层次清楚、专业克制、可落地的研判报告。"
            ),
            max_tokens=3000,
            model="qwen-flash",
        )

    def _generate_document_skill_notes(self, text: str, workflow_key: str) -> str:
        skill_keys = self._workflow_skill_keys(workflow_key)
        skillbook = build_skillbook(skill_keys)
        prompt = (
            "请围绕以下文件，按专家 skills 逐个生成分析笔记。每个 skill 输出一个小节，"
            "格式为 `## 技能名` 加 2 到 3 条短要点。每条要点都必须是具体观察、影响判断或边界提醒。\n\n"
            f"启用技能：\n{skillbook}\n\n"
            f"文件内容：\n{text[:5000]}"
        )
        return self.call_llm(
            prompt,
            system_content="你是金融研究技能编排器，负责先分技能拆解问题，再为后续摘要和报告提供高质量笔记。",
            max_tokens=2200,
            model="qwen-flash",
        )

    def _offline_document_summary(self, text: str) -> str:
        clauses = self._extract_document_clauses(text, limit=4)
        titles = ["政策主旨", "影响路径", "业务观察", "风险提示"]
        lines = []
        for idx, title in enumerate(titles, 1):
            clause = clauses[idx - 1] if idx - 1 < len(clauses) else "原文信息有限，建议结合正式文本复核。"
            lines.append(f"{idx}. **{title}**：{clause}")
        return "\n".join(lines)

    def _offline_document_report(self, text: str) -> str:
        clauses = self._extract_document_clauses(text, limit=6)
        while len(clauses) < 6:
            clauses.append("原文信息有限，当前内容仅作内部研判草稿，正式使用前需结合完整文本复核。")
        return (
            "# 政策概要\n"
            f"{clauses[0]}\n\n"
            "# 关键变化\n"
            f"{clauses[1]}\n\n"
            "# 市场与业务影响\n"
            f"{clauses[2]}\n\n"
            "# 产品策略观察\n"
            f"{clauses[3]}\n\n"
            "# 风险与合规边界\n"
            f"{clauses[4]}\n\n"
            "# 执行优先级\n"
            f"{clauses[5]}"
        )

    def _extract_document_clauses(self, text: str, limit: int = 6) -> List[str]:
        candidates: List[str] = []
        for line in text.splitlines():
            stripped = line.strip()
            if stripped:
                candidates.append(stripped)
        if len(candidates) < limit:
            flattened = re.split(r"[。；;\n]", text)
            for item in flattened:
                stripped = item.strip()
                if stripped:
                    candidates.append(stripped)
        unique: List[str] = []
        seen = set()
        for item in candidates:
            short = item[:120]
            if short in seen:
                continue
            seen.add(short)
            unique.append(short)
            if len(unique) >= limit:
                break
        return unique

    def build_docx_report(self, content: str, title: str = "政策研判报告") -> bytes:
        app_name = self.profile.get("workspace", {}).get("app_name", "资管产品洞察协作台")
        paragraphs = [
            self._docx_paragraph(title, bold=True, size=34),
            self._docx_paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"),
            self._docx_paragraph(f"分析引擎：{app_name}"),
            self._docx_paragraph("=" * 48),
        ]

        for line in content.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            if stripped.startswith("#"):
                heading_text = stripped.replace("#", "").strip()
                paragraphs.append(self._docx_paragraph(heading_text, bold=True, size=26))
            else:
                paragraphs.append(self._docx_paragraph(stripped))

        document_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas" '
            'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
            'xmlns:o="urn:schemas-microsoft-com:office:office" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            'xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math" '
            'xmlns:v="urn:schemas-microsoft-com:vml" '
            'xmlns:wp14="http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing" '
            'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
            'xmlns:w10="urn:schemas-microsoft-com:office:word" '
            'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
            'xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml" '
            'xmlns:wpg="http://schemas.microsoft.com/office/word/2010/wordprocessingGroup" '
            'xmlns:wpi="http://schemas.microsoft.com/office/word/2010/wordprocessingInk" '
            'xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml" '
            'xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape" mc:Ignorable="w14 wp14">'
            f"<w:body>{''.join(paragraphs)}"
            "<w:sectPr><w:pgSz w:w=\"11906\" w:h=\"16838\"/><w:pgMar w:top=\"1440\" w:right=\"1440\" "
            "w:bottom=\"1440\" w:left=\"1440\" w:header=\"708\" w:footer=\"708\" w:gutter=\"0\"/></w:sectPr>"
            "</w:body></w:document>"
        )

        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                "[Content_Types].xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/word/document.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
                "</Types>",
            )
            archive.writestr(
                "_rels/.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
                'Target="word/document.xml"/>'
                "</Relationships>",
            )
            archive.writestr(
                "word/_rels/document.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>',
            )
            archive.writestr("word/document.xml", document_xml)
        buffer.seek(0)
        return buffer.getvalue()

    def _docx_paragraph(self, text: str, bold: bool = False, size: int = 22) -> str:
        safe_text = escape(text)
        bold_xml = "<w:b/>" if bold else ""
        return (
            "<w:p><w:r><w:rPr>"
            f"{bold_xml}<w:sz w:val=\"{size}\"/><w:szCs w:val=\"{size}\"/>"
            "</w:rPr>"
            f"<w:t xml:space=\"preserve\">{safe_text}</w:t>"
            "</w:r></w:p>"
        )

    def _get_llm_api_key(self) -> str:
        return os.getenv("LLM_API_KEY", "").strip() or os.getenv("DASHSCOPE_API_KEY", "").strip()

    def call_llm_stream(
        self,
        prompt: str,
        system_content: str = "你是一个专业的业务助手。",
        history: Optional[List[dict]] = None,
        max_tokens: int = 2000,
        model: str = "qwen-turbo",
    ) -> Iterable[str]:
        api_key = self._get_llm_api_key()
        if not api_key:
            yield self._offline_response(prompt)
            return

        history_str = json.dumps(history or [], ensure_ascii=False)
        state_date = self.get_current_date() or "nodate"
        cache_key = hashlib.md5((model + system_content + prompt + history_str + state_date).encode("utf-8")).hexdigest()
        cached = self._get_cache(cache_key)
        if cached:
            yield cached
            return

        full_response = ""
        try:
            messages = [{"role": "system", "content": system_content}]
            if history:
                messages.extend(history)
            messages.append({"role": "user", "content": prompt})
            payload = json.dumps(
                {
                    "model": model,
                    "messages": messages,
                    "max_tokens": max_tokens,
                    "temperature": 0.5,
                    "stream": True,
                }
            )
            conn = http.client.HTTPSConnection("dashscope.aliyuncs.com", context=ssl.create_default_context())
            headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
            conn.request("POST", "/compatible-mode/v1/chat/completions", payload, headers)
            response = conn.getresponse()
            while True:
                line = response.readline().decode("utf-8").strip()
                if not line:
                    if response.isclosed():
                        break
                    continue
                if not line.startswith("data: "):
                    continue
                data_str = line[6:]
                if data_str == "[DONE]":
                    break
                chunk = json.loads(data_str)
                delta = chunk.get("choices", [{}])[0].get("delta", {})
                content = delta.get("content", "")
                if content:
                    full_response += content
                    yield content
            if full_response:
                self._set_cache(cache_key, prompt, full_response, model)
        except Exception:
            yield self._offline_response(prompt)

    def call_llm(
        self,
        prompt: str,
        system_content: str = "你是一个专业的业务助手。",
        history: Optional[List[dict]] = None,
        max_tokens: int = 2000,
        model: str = "qwen-turbo",
    ) -> str:
        api_key = self._get_llm_api_key()
        if not api_key:
            return self._offline_response(prompt)

        history_str = json.dumps(history or [], ensure_ascii=False)
        state_date = self.get_current_date() or "nodate"
        cache_key = hashlib.md5((model + system_content + prompt + history_str + state_date).encode("utf-8")).hexdigest()
        cached = self._get_cache(cache_key)
        if cached:
            return cached

        try:
            messages = [{"role": "system", "content": system_content}]
            if history:
                messages.extend(history)
            messages.append({"role": "user", "content": prompt})
            payload = json.dumps(
                {
                    "model": model,
                    "messages": messages,
                    "max_tokens": max_tokens,
                    "temperature": 0.5,
                }
            )
            conn = http.client.HTTPSConnection("dashscope.aliyuncs.com", context=ssl.create_default_context())
            headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
            conn.request("POST", "/compatible-mode/v1/chat/completions", payload, headers)
            response = conn.getresponse()
            result = json.loads(response.read().decode("utf-8"))
            content = result.get("choices", [{}])[0].get("message", {}).get("content")
            if not content:
                return self._offline_response(prompt)
            self._set_cache(cache_key, prompt, content, model)
            return content
        except Exception:
            return self._offline_response(prompt)

    def _offline_response(self, prompt: str) -> str:
        trimmed = prompt.strip().replace("\n", " ")
        return (
            "当前未配置语言模型服务，以下为基于现有上下文生成的离线说明：\n\n"
            f"{trimmed[:360]}"
        )

    def get_similar_signals(self, query: str, top_k: int = 3) -> List[dict]:
        if self.vector_store:
            return self.vector_store.search_similar_signals(query, top_k=top_k)
        return []

    def get_user_intent(self, message: str) -> str:
        if self.vector_store:
            return self.vector_store.get_intent(message)
        if any(keyword in message for keyword in ["政策", "监管", "文件", "合规"]):
            return "policy"
        if any(keyword in message for keyword in ["建议", "策略", "传播", "文案"]):
            return "planning"
        if any(keyword in message for keyword in ["对比", "趋势", "变化", "区间"]):
            return "comparison"
        return "query"
