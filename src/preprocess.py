from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent


def merge_dicts(base: dict, override: dict) -> dict:
    merged = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = merge_dicts(merged[key], value)
        else:
            merged[key] = value
    return merged


def load_profile(profile_path: Optional[Path]) -> dict:
    default_path = BASE_DIR / "config" / "default_profile.json"
    profile = json.loads(default_path.read_text(encoding="utf-8"))
    if profile_path and profile_path.exists():
        custom = json.loads(profile_path.read_text(encoding="utf-8"))
        profile = merge_dicts(profile, custom)
    return profile


def normalize_date(value: object) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def analyze_snapshot(data_dir: Path, profile: dict) -> dict:
    snapshot_profile = profile.get("snapshot", {})
    files = profile.get("files", {})
    snapshot_path = data_dir / files.get("market_snapshot", "market_snapshot.json")
    report = {
        "exists": snapshot_path.exists(),
        "path": str(snapshot_path),
        "date_count": 0,
        "record_count": 0,
        "index_count": 0,
        "invalid_dates": 0,
        "missing_name": 0,
        "missing_code": 0,
        "missing_index_code": 0,
    }
    if not snapshot_path.exists():
        return report

    raw = json.loads(snapshot_path.read_text(encoding="utf-8"))
    date_field = snapshot_profile.get("date_field", "date")
    products_field = snapshot_profile.get("products_field", "products")
    indices_field = snapshot_profile.get("indices_field", "indices")
    product_fields = snapshot_profile.get("product_fields", {})
    index_fields = snapshot_profile.get("index_fields", {})

    if isinstance(raw, dict):
        iterable = raw.items()
    elif isinstance(raw, list):
        iterable = [(item.get(date_field), item) for item in raw if isinstance(item, dict)]
    else:
        iterable = []

    for raw_date, payload in iterable:
        if not isinstance(payload, dict) or not normalize_date(raw_date):
            report["invalid_dates"] += 1
            continue
        report["date_count"] += 1
        products = payload.get(products_field, [])
        indices = payload.get(indices_field, {})
        if isinstance(products, list):
            report["record_count"] += len(products)
            for item in products:
                if not isinstance(item, dict):
                    continue
                if not item.get(product_fields.get("name", "name")):
                    report["missing_name"] += 1
                if not item.get(product_fields.get("code", "code")):
                    report["missing_code"] += 1
                if not item.get(product_fields.get("index_code", "index_code")):
                    report["missing_index_code"] += 1
        if isinstance(indices, dict):
            report["index_count"] += len(indices)
        elif isinstance(indices, list):
            report["index_count"] += len(
                [item for item in indices if isinstance(item, dict) and item.get(index_fields.get("code", "code"))]
            )
    report["missing_index_ratio"] = round(report["missing_index_code"] / report["record_count"], 2) if report["record_count"] else 0.0
    return report


def match_header(header_lookup: Dict[str, int], aliases: List[str]) -> Optional[int]:
    lowered = {key.lower(): value for key, value in header_lookup.items()}
    for alias in aliases:
        if alias in header_lookup:
            return header_lookup[alias]
        if alias.lower() in lowered:
            return lowered[alias.lower()]
    return None


def analyze_policy_catalog(data_dir: Path, profile: dict) -> dict:
    policy_profile = profile.get("policy", {})
    files = profile.get("files", {})
    policy_path = data_dir / files.get("policy_catalog", "policy_catalog.xlsx")
    report = {
        "exists": policy_path.exists(),
        "path": str(policy_path),
        "row_count": 0,
        "invalid_dates": 0,
        "missing_titles": 0,
        "resolved_headers": {},
    }
    if not policy_path.exists():
        return report

    workbook = load_workbook(policy_path, data_only=True, read_only=True)
    sheet_name = policy_profile.get("sheet_name")
    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
    headers = [cell.value for cell in sheet[1]]
    header_lookup = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}
    columns = policy_profile.get("columns", {})
    date_idx = match_header(header_lookup, columns.get("date", ["公告日期"]))
    title_idx = match_header(header_lookup, columns.get("title", ["标题"]))
    rank_idx = match_header(header_lookup, columns.get("rank", ["法律位阶"]))
    source_idx = match_header(header_lookup, columns.get("source", ["来源"]))
    report["resolved_headers"] = {
        "date": date_idx is not None,
        "title": title_idx is not None,
        "rank": rank_idx is not None,
        "source": source_idx is not None,
    }

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        report["row_count"] += 1
        values = list(row)
        raw_date = values[date_idx] if date_idx is not None and date_idx < len(values) else None
        raw_title = values[title_idx] if title_idx is not None and title_idx < len(values) else None
        if not normalize_date(raw_date):
            report["invalid_dates"] += 1
        if not str(raw_title or "").strip():
            report["missing_titles"] += 1
    return report


def validate_data_dir(data_dir: Path, profile_path: Optional[Path] = None) -> dict:
    profile = load_profile(profile_path)
    snapshot = analyze_snapshot(data_dir, profile)
    policy = analyze_policy_catalog(data_dir, profile)
    ready = snapshot["exists"] and policy["exists"] and snapshot["record_count"] > 0 and policy["row_count"] > 0
    return {
        "data_dir": str(data_dir),
        "profile_path": str(profile_path) if profile_path else str(BASE_DIR / "config" / "default_profile.json"),
        "ready": ready,
        "snapshot": snapshot,
        "policy_catalog": policy,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate an external data directory with profile-aware rules.")
    parser.add_argument("data_dir", type=Path, help="Directory containing the configured snapshot and policy files")
    parser.add_argument("--profile", type=Path, default=None, help="Optional DATA_PROFILE_PATH override")
    args = parser.parse_args()
    report = validate_data_dir(args.data_dir, args.profile)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
